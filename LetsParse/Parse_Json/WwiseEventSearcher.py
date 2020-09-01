import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from waapi import WaapiClient, CannotConnectToWaapiException
import os
from jsonpath_ng import parse
from collections import defaultdict
from PySide2.QtWidgets import QWidget, QVBoxLayout, QMessageBox, QHBoxLayout, QTableWidget, QLineEdit, QPushButton, QApplication, QLabel, QTableWidgetItem, QFileDialog, QAbstractItemView



#parents = '..\\Assets\\Resources\\Outgame\\Data\\Sound' # 실제 path
parents = os.getcwd()  # 개발전용
os.chdir(parents)  # json 폴더 지정


header = ['file', 'ContentsKey', 'animName', 'soundName', 'targetObjName', 'sequenceTime', 'playOneShot', 'dontDestroy']

def jsontodict(sound):
    results = defaultdict(dict)
    i = 0

    for root, dirs, files in os.walk('./'):
        for file in files:
            if file.lower().endswith('.json'):  # 특정 확장자만 열기
                fpath = os.path.join(root, file)  # join 으로 합쳐야지 str이 아닌 dir로 인식
                with open(fpath, encoding='cp949') as json_file:  # 인코딩 지정하지 않으면 에러 발
                    json_data = json.load(json_file)
                    json_str = json.dumps(json_data)

                    for key in parse('$..ContentsKey').find(json.loads(json_str)):
                        data = key.context.value['m_saveDataList']
                        for match in parse('$..soundName').find(data):
                            if sound == '':  # 빈 칸이면 모두
                                results[i][header[0]] = file
                                results[i][header[1]] = key.context.value[header[1]]
                                results[i].update(match.context.value)
                                i += 1
                            elif sound.lower() in match.value.lower():  # 특정 단어 포함하는지, 대소문자 구분 없이 하기 위해 둘다 소문자 처리해서 비교
                                results[i][header[0]] = file
                                results[i][header[1]] = key.context.value[header[1]]
                                results[i].update(match.context.value)
                                i += 1

    return results


class Form(QWidget):
    def __init__(self):
        super(Form, self).__init__()
        self.setWindowTitle("SearchSound")
        self.setMinimumSize(875, 400)

        self.vb = QVBoxLayout()
        self.setLayout(self.vb)
        self.hbTop = QHBoxLayout()
        self.hbMid = QVBoxLayout()
        self.hbMidBot = QHBoxLayout()
        self.hbBot = QHBoxLayout()
        self.vb.addLayout(self.hbTop)
        self.vb.addLayout(self.hbMid)
        self.vb.addLayout(self.hbMidBot)
        self.vb.addLayout(self.hbBot)

        self.ln = QLineEdit("검색 할 Event 명에 들어간 단어를 입력해 주세요")
        self.btn_name = QPushButton("Search")
        self.btn_all = QPushButton("Search All")
        self.btn_open = QPushButton("Open Json Path")
        self.btn_play = QPushButton("Play Sound")
        self.lb_result = QLabel("결과 :")
        self.tb_result = QTableWidget()
        self.tb_result.setAutoScroll(True)
        self.tb_result.showGrid()

        self.btn_save = QPushButton("Save As Xlsx")
        self.message = QMessageBox()

        self.hbTop.addWidget(self.ln)
        self.hbTop.addWidget(self.btn_name)
        self.hbTop.addWidget(self.btn_all)
        self.hbMid.addWidget(self.lb_result)
        self.hbMid.addWidget(self.tb_result)
        self.hbMidBot.addWidget(self.btn_open)
        self.hbMidBot.addWidget(self.btn_play)
        self.hbBot.addWidget(self.btn_save)

        self.btn_name.clicked.connect(self.search)
        self.btn_all.clicked.connect(self.search_all)
        self.btn_save.clicked.connect(self.savefile)

        self.btn_play.clicked.connect(self.play)
        self.btn_open.clicked.connect(self.open)


    def search_all(self):
        self.showresult(jsontodict(''))
        return

    def search(self):
        self.showresult(jsontodict(self.ln.text()))
        return

    def play(self):
        self.playsound()
        return

    def open(self):
        self.openjson()
        return

    def playsound(self):
        try:
            client = WaapiClient()

            langargs = {
                "event": self.tb_result.item(self.tb_result.currentRow(), 3).text(),
                "gameObject": 0
            }

            client.call("ak.wwise.ui.bringToForeground")
            client.call("ak.soundengine.postEvent", langargs)

        except CannotConnectToWaapiException:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("Search Sound")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("WAAPI에 연결하지 못했습니다. : Wwise가 켜져있고 WAAPI가 Enabled 되어 있는지 체크 해주세요.")
            self.message0.exec()

    def openjson(self): #아무 응답이 없을 때는 연결프로그램 확인
        os.startfile(parents + '/' + self.tb_result.item(self.tb_result.currentRow(), 0).text())
        return


    def showresult(self, dict):

        self.tb_result.clear() # 채우기 전에 초기화

        self.tb_result.setRowCount(len(dict))
        self.lb_result.setText("결과 : 총 " + str(len(dict)) + " 개의 사운드를 찾았습니다.")
        self.repaint() # 이걸 해줘야 레이블이 업데이트 됨

        self.tb_result.setColumnCount(len(header))
        self.tb_result.setHorizontalHeaderLabels(header)

        for i in dict:
            for k in range(len(header)):
                item = QTableWidgetItem(str(dict[i][header[k]]))
                self.tb_result.setItem(i, k, item)

        # len(dict) == i 는 row, 행의 갯수
        # len(dict[0]) == len(header) == k 는 col, 열의 갯수

        self.tb_result.setEditTriggers(QTableWidget.NoEditTriggers) # 에디팅 막음
        self.tb_result.setSelectionMode(QAbstractItemView.SingleSelection) # 중복선택 불가능 하게

        return

    def savefile(self):
        filename = QFileDialog.getSaveFileName(self, 'Save file', "", ".xlsx(*.xlsx)")

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sound'

        rowindex = 1
        colindex = 1

        for h in header:
            ws.cell(row=rowindex, column=colindex).value = h
            ws.cell(row=rowindex, column=colindex).font = Font(bold=True, color='ffffff')
            ws.cell(row=rowindex, column=colindex).fill = PatternFill("solid", fgColor="404040")
            colindex += 1

        ws.freeze_panes = 'A2'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30

        ws.auto_filter.ref = "A1:H1"


        for currentColumn in range(1, self.tb_result.columnCount()+1):
            for currentRow in range(1, self.tb_result.rowCount()+1):
                try:
                    teext = str(self.tb_result.item(currentRow-1, currentColumn-1).text())
                    ws.cell(currentRow+1, currentColumn).value = teext

                except AttributeError:
                    pass



        wb.save(filename[0])

        self.message1 = QMessageBox()
        self.message1.setWindowTitle("Search Sound")
        self.message1.setText(str(filename[0]) + '\n' + '\n' + "저장 완료!")
        self.message1.exec()

        wb.close()

        return

app = QApplication([])
GUI = Form()

GUI.show()
app.exec_()
