import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os

parents = '..\\Assets\\Resources\\Outgame\\Data\\Sound'
os.chdir(parents) # json 폴더 지정

'''
Json --> Dictionary
'''
def ParseJson():
    all = []
    alldict = {}

    for root, dirs, files in os.walk('./'):
        for file in files:
            if file.lower().endswith('.json'):  # 특정 확장자만 열기
                fpath = os.path.join(root, file)  # join 으로 합쳐야지 str이 아닌 dir로 인식
                with open(fpath, encoding='cp949') as json_file:
                    json_data = json.load(json_file)
                    json_str = json.dumps(json_data)
                    json_dict = json.loads(json_str)
                    all.append(json_dict)

    for i in range(len(all)):
        if 'm_saveDataList' in all[i]:
            pass
            #alldict['CharacterName'] = all[0]["ContentsKey"]
            #print(all[i]['m_saveDataList'])
            for n in range(len(all[i]['m_saveDataList'])):
                #print(all[i]['m_saveDataList'][n]['EventSoundDataList'])
                for k in range(len(all[i]['m_saveDataList'][n]['EventSoundDataList'])):
                    print (all[i]['m_saveDataList'][n]['EventSoundDataList'][k])
        elif 'ObjectDataList' in all[i]:
            #alldict['ThemeName'] = all[0]["ThemeKey"]
            #print(all[i]['ObjectDataList'])
            for o in range(len(all[i]['ObjectDataList'])):
                for m in range(len(all[i]['ObjectDataList'][o]['m_saveDataList'])):
                    for h in range(len(all[i]['ObjectDataList'][o]['m_saveDataList'][m]['EventSoundDataList'])):
                        print (all[i]['ObjectDataList'][o]['m_saveDataList'][m]['EventSoundDataList'][h])
    return print(alldict)

ParseJson()
#
# all = [] # WAAPI와도 통신 하기 위해서 따로 빼놓기
# for i in range(len(json_dict['datas'])):
#     for k in json_dict['datas'][i].keys():  # k = theme, 캐릭터 이름
#         dict1 = {}
#         dict1['Theme Or Character'] = k
#         for s in range(len(json_dict['datas'][i][k])):
#             for n in json_dict['datas'][i][k][s].keys():  # n = 애니메인션 이름
#                 for r in range(len(json_dict['datas'][i][k][s][n])):
#                     dict2 = json_dict['datas'][i][k][s][n][r]
#                     dict3 = {**dict1, **dict2} # 딕셔너리 더하기
#                     all.append(dict3)
#
#
# print(all)
# '''
# 엑셀로 쓰는 함수, 첫번째 행은 key들을 넣어주고 두번째 행부터는 value들
# '''
#
# wb = Workbook()
# ws = wb.active
# ws.title = 'EventSound'
#
# def toxls(all):
#
#     rowindex = 1
#     colindex = 1
#
#     for key in all[0]:
#         ws.cell(row=rowindex, column=colindex).value = key
#         ws.cell(row=rowindex, column=colindex).font = Font(bold=True, color='ffffff')
#         ws.cell(row=rowindex, column=colindex).fill = PatternFill("solid", fgColor="404040")
#         colindex += 1
#
#     ws.freeze_panes = 'A2'
#     ws.column_dimensions['A'].width = 30
#     ws.column_dimensions['B'].width = 30
#     ws.column_dimensions['C'].width = 30
#     ws.column_dimensions['D'].width = 30
#     ws.column_dimensions['E'].width = 30
#
#     ws.auto_filter.ref = "A1:E1"
#
#     colindex = 1
#     rowindex = 2
#
#     for dict in all:
#         for data in dict:
#             ws.cell(row=rowindex, column=colindex).value = dict[data]
#             colindex += 1
#         rowindex += 1
#         colindex = 1
#
#
# toxls(all)
# wb.save('./EventSound.xlsx')
# wb.close()
#
#
#
# '''
# GUI 코드
# # '''
# from PySide2.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox, QMessageBox, QApplication
#
#
# class Form(QWidget):
#     def __init__(self):
#         super(Form, self).__init__()
#         self.setWindowTitle("JSON To XLSX")
#         self.setMinimumSize(350, 200)
#         self.setMaximumSize(350, 200)
#
#
#
# app = QApplication([])
# GUI = Form()
#
# GUI.show()
# app.exec_()
