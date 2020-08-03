import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os
from jsonpath_ng import parse



def searchReference(sound):
    # parents = '..\\Assets\\Resources\\Outgame\\Data\\Sound' # Windows
    parents = '../Assets/Resources/Outgame/Data/Sound'  # MacOS
    os.chdir(parents)  # json 폴더 지정

    results = []  # json 이름, Contents Key, 노티파이정보'

    for root, dirs, files in os.walk('./'):
        for file in files:
            if file.lower().endswith('.json'):  # 특정 확장자만 열기
                fpath = os.path.join(root, file)  # join 으로 합쳐야지 str이 아닌 dir로 인식
                with open(fpath, encoding='cp949') as json_file: # 인코딩 지정하지 않으면 에러 발
                    json_data = json.load(json_file)
                    json_str = json.dumps(json_data)
                    json_dict = json.loads(json_str)

                searchsoundName = parse('$..soundName') # '$..EventSoundDataList.[*].soundName'와 같음
                searchContentsKey = parse('$..ContentsKey')

                for match in searchsoundName.find(json_dict):
                    results.append(file)
                    results.append(searchContentsKey.find(json_dict)[0].value)
                    results.append(match.context.value)
    return results

def allsoundtoxlsx(jsons):
    results = []  # json 이름, Contents Key, 노티파이정보'
    result = {}

    for json in jsons:
        fpath = os.path.join(json)
        print(fpath)
        with open(fpath, encoding='cp949') as json_file:  # 인코딩 지정하지 않으면 에러 발
            json_data = json.load(json_file)
            json_str = json.dumps(json_data)
            json_dict = json.loads(json_str)

        searchsoundName = parse('$..')

        for match in searchsoundName.find(json_dict):
            print(match.value)

    return results

print(allsoundtoxlsx(['../Assets/Resources/Outgame/Data/Sound/Theme/theme_1.json'])) #절대경로여야할듯

#
#
# '''
# 엑셀로 쓰는 함수, 첫번째 행은 key들을 넣어주고 두번째 행부터는 value들
# '''
#
# wb = Workbook()
# ws = wb.active
# ws.title = 'EventSound'
#
# def toxls():
#
#     rowindex = 1
#     colindex = 1
#
#     for key in ParseJson()[0]:
#         ws.cell(row=rowindex, column=colindex).value = key
#         ws.cell(row=rowindex, column=colindex).font = Font(bold=True, color='ffffff')
#         ws.cell(row=rowindex, column=colindex).fill = PatternFill("solid", fgColor="404040")
#         colindex += 1
#
#     ws.freeze_panes = 'A2'
#     ws.column_dimensions['A'].width = 30
#     ws.column_dimensions['B'].width = 30
#     ws.column_dimensions['C'].width = 30
#     ws.column_dimensions['D'].width = 30
#     ws.column_dimensions['E'].width = 30
#
#     ws.auto_filter.ref = "A1:G1"
#
#     colindex = 1
#     rowindex = 2
#
#     for dict in ParseJson():
#         for data in dict:
#             ws.cell(row=rowindex, column=colindex).value = dict[data]
#             colindex += 1
#         rowindex += 1
#         colindex = 1
#
#
# toxls()
# wb.save('C:\\Users\\trippysour\\Desktop\\EventSound.xlsx')
# wb.close()
#

#
# '''
# GUI 코드
# # '''
# from PySide2.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox, QMessageBox, QApplication
#
# class Form(QWidget):
#     def __init__(self):
#         super(Form, self).__init__()
#         self.setWindowTitle("JSON To XLSX")
#         self.setMinimumSize(350, 200)
#         self.setMaximumSize(350, 200)
#
# #결과 팝업창, json 열기 버튼, exel 시트 만들기버튼
#
# app = QApplication([])
# GUI = Form()
#
# GUI.show()
# app.exec_()

