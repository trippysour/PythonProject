import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import xml.etree.ElementTree as ET
from collections import defaultdict
from PySide2.QtWidgets import QWidget, QGroupBox, QTabWidget, QCheckBox, QMessageBox, QHBoxLayout, QVBoxLayout, QTableWidget, QLineEdit, QPushButton, QApplication, QLabel, QTableWidgetItem, QFileDialog, QAbstractItemView
from PySide2.QtGui import QColor


#parents = '..\\Assets\\Resources\\Outgame\\Data\\Sound' # 실제 path
parents = os.getcwd()  # 개발전용
os.chdir(parents)  # json 폴더 지정


def sound_in_anim(name):
    header = ['xml', 'seq_name', 'file', 'TOOL_InRadius', 'TOOL_OutRadius', 'vol']
    results = defaultdict(dict)
    i = 0

    for root, dirs, files in os.walk('./'):
        for file in files:
            if file.lower().endswith('.xml'):  # 특정 확장자만 열기
                xml = ET.parse(file)
                root = xml.getroot()

                for seq in root.iter('animation'):
                    for sounds in seq.findall('sound'):
                        if name == '':
                            if sounds.get('file') != None:
                                results[i][header[0]] = file
                                results[i][header[1]] = seq.get('name')
                                for k in range(2, 6):
                                    results[i][header[k]] = sounds.get(header[k])
                                i += 1
                            else:
                                for sound in sounds.findall('probsound'):
                                    results[i][header[0]] = file
                                    results[i][header[1]] = seq.get('name')
                                    for k in range(2, 6):
                                        results[i][header[k]] = sound.get(header[k])
                                    i += 1
                        else:
                            if sounds.get('file') != None:
                                if name.lower() in sounds.get('file').lower():
                                    results[i][header[0]] = file
                                    results[i][header[1]] = seq.get('name')
                                    for k in range(2, 6):
                                        results[i][header[k]] = sounds.get(header[k])
                                    i += 1
                            else:
                                for sound in sounds.findall('probsound'):
                                    if name.lower() in sound.get('file').lower():
                                        results[i][header[0]] = file
                                        results[i][header[1]] = seq.get('name')
                                        for k in range(2, 6):
                                            results[i][header[k]] = sound.get(header[k])
                                        i += 1
    return results, header


def sound_in_particles(name):
    header = ['file', 'particle_name', 'Sound', 'SoundMinRange', 'SoundMaxRange', 'SoundVolume', 'SoundLoop']
    results = defaultdict(dict)
    i = 0

    for root, dirs, files in os.walk('./'):
        for file in files:
            if file.lower().endswith('.xml'):  # 특정 확장자만 열기
                xml = ET.parse(file)
                root = xml.getroot()

                for particles in root.iter('Particles'):
                    for sounds in particles:
                        if name == '':
                            if sounds.get('Sound') != None and sounds.get('Sound') != '':
                                results[i][header[0]] = file
                                results[i][header[1]] = particles.get('Name')
                                for k in range(2, 7):
                                    results[i][header[k]] = sounds.get(header[k])
                                i += 1
                        else:
                            if sounds.get('Sound') != None and name.lower() in sounds.get('Sound').lower():
                                results[i][header[0]] = file
                                results[i][header[1]] = particles.get('Name')
                                for k in range(2, 7):
                                    results[i][header[k]] = sounds.get(header[k])
                                i += 1
    return results, header


def sound_in_lyr(sound):

    header_SoundSpot = ['lyr', 'Layer', 'Name', 'Pos', 'EventType', 'bLoop', 'bOnce', 'bPlay', 'sndSource', 'InnerRadius', 'OuterRadius', 'iVolume']
    header_RA = ['lyr', 'Layer', 'Name', 'EventType', 'sndSound', 'bCentered', 'iVolume', 'iChanceOfOccuring', 'bDoNotOverlap']
    header_Shape = ['lyr', 'Layer', 'Name']

    results_SoundSpot = defaultdict(dict)
    results_RA = defaultdict(dict)
    results_Shape = defaultdict(dict)
    i = 0
    s = 0
    r = 0

    for root, fir, files in os.walk('./'):
        for file in files:
            if file.lower().endswith('.lyr'):
                xml = ET.parse(file)
                root = xml.getroot()

                for object in root.iter('Object'):

                    if object.get('Type') == 'SoundSpot':
                        if sound == '':
                            results_SoundSpot[s][header_SoundSpot[0]] = file
                            for k in range(1, 5):
                                results_SoundSpot[s][header_SoundSpot[k]] = object.get(header_SoundSpot[k])
                            for k in range(5, 12):
                                results_SoundSpot[s][header_SoundSpot[k]] = object.find('Properties').get(
                                    header_SoundSpot[k])
                            s += 1
                        elif sound.lower() in object.find('Properties').get('sndSource').lower():
                            results_SoundSpot[s][header_SoundSpot[0]] = file
                            for k in range(1, 5):
                                results_SoundSpot[s][header_SoundSpot[k]] = object.get(header_SoundSpot[k])
                            for k in range(5, 12):
                                results_SoundSpot[s][header_SoundSpot[k]] = object.find('Properties').get(
                                    header_SoundSpot[k])
                            s += 1

                    elif object.get('Type') == 'RandomAmbient':
                        if sound == '':
                            for v in object.iter():
                                if v.tag.startswith('Sound') and v.attrib['sndSound'] != '':
                                    results_RA[r][header_RA[0]] = file
                                    for k in range(1, 4):
                                        results_RA[r][header_RA[k]] = object.get(header_RA[k])
                                        results_RA[r].update(v.attrib)
                                    r += 1
                        else:
                            for v in object.iter():
                                if v.tag.startswith('Sound') and sound.lower() in v.attrib['sndSound'].lower():
                                    results_RA[r][header_RA[0]] = file
                                    for k in range(1, 4):
                                        results_RA[r][header_RA[k]] = object.get(header_RA[k])
                                        results_RA[r].update(v.attrib)
                                    r += 1

                    # elif object.get('Type') == 'Shape':
                    #     if sound == '':
                    #         results_Shape[i][header_Shape[0]] = file
                    #         for k in range(1, 3):
                    #             results_Shape[i][header_Shape[k]] = object.get(header_Shape[k])
                    #         i += 1
                    #     else:
                    #         pass

    return [results_SoundSpot, header_SoundSpot], [results_RA, header_RA], [results_Shape, header_Shape]

#print(sound_in_lyr('')[2])

class Form(QWidget):
    def __init__(self):
        super(Form, self).__init__()
        self.setWindowTitle("SoundSearch_AION")
        self.setMinimumSize(1000, 800)
        '''
        Layout_Base
        '''
        self.vb = QVBoxLayout()
        self.setLayout(self.vb)

        self.hbTop = QHBoxLayout()
        self.hbMid = QVBoxLayout()
        self.hbMidBot = QHBoxLayout()
        self.hbBot = QHBoxLayout()


        '''
        Groubbox
        '''
        self.gb_data = QGroupBox('Data Type')
        self.CB_Anim = QCheckBox("AnimationMarkers")
        self.CB_Particle = QCheckBox("Particles")
        self.CB_Layers = QCheckBox("Layers")
        self.CB_Anim.setChecked(True)
        self.CB_Particle.setChecked(True)
        self.CB_Layers.setChecked(True)
        self.ln = QLineEdit("")
        self.btn_name = QPushButton("Search")

        self.vbox = QHBoxLayout()
        self.vbox.addWidget(self.CB_Anim)
        self.vbox.addWidget(self.CB_Particle)
        self.vbox.addWidget(self.CB_Layers)
        self.vbox.addWidget(self.ln)
        self.vbox.addWidget(self.btn_name)
        self.gb_data.setLayout(self.vbox)
        self.hbTop.addWidget(self.gb_data)

        self.vb.addLayout(self.hbTop)
        self.vb.addLayout(self.hbMid)
        self.vb.addLayout(self.hbMidBot)
        self.vb.addLayout(self.hbBot)

        '''
        Button
        '''

        self.btn_open = QPushButton("Open Selected Data File")
        self.btn_play = QPushButton("Open Selected Sound")

        self.btn_save = QPushButton("Save And Open xlsx")
        self.message = QMessageBox()

        self.hbMidBot.addWidget(self.btn_open)
        self.hbMidBot.addWidget(self.btn_play)
        self.hbBot.addWidget(self.btn_save)

        self.ln.returnPressed.connect(self.search)
        self.btn_name.clicked.connect(self.search)
        self.btn_save.clicked.connect(self.savefile)
        self.btn_play.clicked.connect(self.play)
        self.btn_open.clicked.connect(self.open)

        '''
        result
        '''
        self.tabs = QTabWidget()
        self.hbMid.addWidget(self.tabs)
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, 'Result')



    def addTab(self, name, dicts):
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, name)

        self.tab1.layout = QVBoxLayout(self)
        self.tab1.setLayout(self.tab1.layout)

        dict = dicts[0]
        header = dicts[1]

        self.tb_result = QTableWidget()
        self.tb_result.setAutoScroll(True)
        self.tb_result.showGrid()
        self.tb_result.clear()  # 채우기 전에 초기화

        self.tb_result.setRowCount(len(dict))
        self.tb_result.setMinimumHeight(200)
        self.repaint()  # 이걸 해줘야 레이블이 업데이트 됨

        self.tb_result.setColumnCount(len(header))
        self.tb_result.setHorizontalHeaderLabels(header)

        for i in dict:
            for k in range(len(header)):
                item = QTableWidgetItem(str(dict[i][header[k]]))
                self.tb_result.setItem(i, k, item)

        # len(dict) == i 는 row, 행의 갯수
        # len(dict[0]) == len(header) == k 는 col, 열의 갯수

        self.tb_result.setEditTriggers(QTableWidget.NoEditTriggers)  # 에디팅 막음
        for i in range(len(header)):
            self.tb_result.setColumnWidth(i, 905 // len(header))
        self.tb_result.setSelectionMode(QAbstractItemView.SingleSelection)  # 중복선택 불가능 하게
        self.tab1.layout.addWidget(self.tb_result)


    def search(self):

        self.tabs.clear()

        if self.CB_Anim.isChecked() and len(sound_in_anim(self.ln.text())[0]) != 0:
            self.addTab('AnimationMarkers', sound_in_anim(self.ln.text()))

        if self.CB_Particle.isChecked() and len(sound_in_particles(self.ln.text())[0]) != 0:
            self.addTab('Particles', sound_in_particles(self.ln.text()))

        if self.CB_Layers.isChecked():
            if len(sound_in_lyr(self.ln.text())[0][0]) != 0:
                self.addTab('Layer_Spot', sound_in_lyr(self.ln.text())[0])
            if len(sound_in_lyr(self.ln.text())[1][0]) != 0:
                self.addTab('Layer_RA', sound_in_lyr(self.ln.text())[1])

        return

    def play(self):
        self.postevent()
        return

    def open(self):
        self.openjson()
        return




    def openjson(self): #아무 응답이 없을 때는 연결프로그램 확인

        print(self.tabs.children())


        try:
            os.startfile(parents + '/' + self.tb_result.item(self.tb_result.currentRow(), 0).text())

        except AttributeError:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_AION")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("실행할 Json을 선택해 주세요.")
            self.message0.exec()
        return

    def savefile(self, header):
        filename = QFileDialog.getSaveFileName(self, 'Save file', "", ".xlsx(*.xlsx)")

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sound'

        rowindex = 1
        colindex = 1

        for h in header:
            ws.cell(row=rowindex, column=colindex).value = h
            ws.cell(row=rowindex, column=colindex).font = Font(bold=True, color='ffffff')
            ws.cell(row=rowindex, column=colindex).fill = PatternFill("solid", fgColor="404040")
            colindex += 1

        ws.freeze_panes = 'A2'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30

        ws.auto_filter.ref = "A1:H1"


        for currentColumn in range(1, self.tb_result.columnCount()+1):
            for currentRow in range(1, self.tb_result.rowCount()+1):
                try:
                    teext = str(self.tb_result.item(currentRow-1, currentColumn-1).text())
                    ws.cell(currentRow+1, currentColumn).value = teext

                except AttributeError:
                    pass


        try:
            wb.save(filename[0])
            self.message1 = QMessageBox()
            self.message1.setWindowTitle("SoundSearch_AION")
            self.message1.setText(str(filename[0]) + '\n' + '\n' + "저장 완료!")
            self.message1.exec()
            os.startfile(str(filename[0]))
        except PermissionError:
            self.message1 = QMessageBox()
            self.message1.setWindowTitle("SoundSearch_AION")
            self.message1.setText("저장 권한이 없습니다. 엑셀이 닫혀있는지 확인해 주세요.")
            self.message1.exec()




        wb.close()

        return


app = QApplication([])
GUI = Form()
GUI.show()
app.exec_()
