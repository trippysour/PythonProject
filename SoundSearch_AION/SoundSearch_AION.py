import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.cell import get_column_letter
import xml.etree.ElementTree as ET
from collections import defaultdict
from PySide2.QtWidgets import QListWidgetItem, QListWidget, QWidget, QGroupBox, QTabWidget, QCheckBox, QMessageBox, QHBoxLayout, QVBoxLayout, QTableWidget, QLineEdit, QPushButton, QApplication, QTableWidgetItem, QFileDialog, QAbstractItemView
from PySide2 import QtCore

dropedlist = []

class Drop:
    def append(self, file):
        global dropedlist
        dropedlist.append(file)
    def remove(self):
        global dropedlist
        dropedlist = []


classfier = Drop()


def sound_in_anim(name):

    header = ['data', 'seq_name', 'file', 'TOOL_InRadius', 'TOOL_OutRadius', 'vol']
    results = defaultdict(dict)
    i = 0

    if len(dropedlist) < 1:

        parents = os.path.abspath(os.path.join(os.path.abspath(os.pardir), os.pardir))
        current = 'AION_Data//Data//animationmarkers'
        path = os.path.join(parents, current)

        for root, dirs, files in os.walk(path):

            for file in files:

                if file.lower().endswith('.xml'):
                    xml = ET.parse(os.path.join(path, file))
                    root = xml.getroot()
                    print('AnimationMarkers - ' + file + ' 탐색중...')

                    for seq in root.iter('animation'):
                        for sounds in seq.findall('sound'):
                            if name == '':
                                if sounds.get('file') != None:
                                    results[i][header[0]] = file
                                    results[i][header[1]] = seq.get('name')
                                    for k in range(2, 6):
                                        results[i][header[k]] = sounds.get(header[k])
                                    i += 1
                                else:
                                    for sound in sounds.findall('probsound'):
                                        results[i][header[0]] = file
                                        results[i][header[1]] = seq.get('name')
                                        for k in range(2, 6):
                                            results[i][header[k]] = sound.get(header[k])
                                        i += 1
                            else:
                                if sounds.get('file') != None:
                                    if name.lower() in sounds.get('file').split('\\')[-1].lower():
                                        results[i][header[0]] = file
                                        results[i][header[1]] = seq.get('name')
                                        for k in range(2, 6):
                                            results[i][header[k]] = sounds.get(header[k])
                                        i += 1
                                else:
                                    for sound in sounds.findall('probsound'):
                                        if name.lower() in sound.get('file').split('\\')[-1].lower():
                                            results[i][header[0]] = file
                                            results[i][header[1]] = seq.get('name')
                                            for k in range(2, 6):
                                                results[i][header[k]] = sound.get(header[k])
                                            i += 1
    else:
        path = ''

        for file in dropedlist:
            if file.lower().endswith('.xml'):
                xml = ET.parse(os.path.join(path, file))
                root = xml.getroot()
                try:
                    if root.iter('animation') is not None: print('AnimationMarkers - ' + file + ' 탐색중...')
                except:
                    pass

                for seq in root.iter('animation'):
                    for sounds in seq.findall('sound'):
                        if name == '':
                            if sounds.get('file') != None:
                                results[i][header[0]] = file
                                results[i][header[1]] = seq.get('name')
                                for k in range(2, 6):
                                    results[i][header[k]] = sounds.get(header[k])
                                i += 1
                            else:
                                for sound in sounds.findall('probsound'):
                                    results[i][header[0]] = file
                                    results[i][header[1]] = seq.get('name')
                                    for k in range(2, 6):
                                        results[i][header[k]] = sound.get(header[k])
                                    i += 1
                        else:
                            if sounds.get('file') != None:
                                if name.lower() in sounds.get('file').split('\\')[-1].lower():
                                    results[i][header[0]] = file
                                    results[i][header[1]] = seq.get('name')
                                    for k in range(2, 6):
                                        results[i][header[k]] = sounds.get(header[k])
                                    i += 1
                            else:
                                for sound in sounds.findall('probsound'):
                                    if name.lower() in sound.get('file').split('\\')[-1].lower():
                                        results[i][header[0]] = file
                                        results[i][header[1]] = seq.get('name')
                                        for k in range(2, 6):
                                            results[i][header[k]] = sound.get(header[k])
                                        i += 1



    return results, header


def sound_in_particles(name):
    header = ['data', 'particle_name', 'Sound', 'SoundMinRange', 'SoundMaxRange', 'SoundVolume', 'SoundLoop']
    results = defaultdict(dict)
    i = 0

    if len(dropedlist) < 1:

        parents = os.path.abspath(os.path.join(os.path.abspath(os.pardir), os.pardir))
        current = 'AION_Data//Editor//particles'
        path = os.path.join(parents, current)

        for root, dirs, files in os.walk(path):
            for file in files:
                if file.lower().endswith('.xml'):
                    print('Particles - ' + file + ' 탐색중...')
                    xml = ET.parse(os.path.join(path, file))
                    root = xml.getroot()

                    for particles in root.iter('Particles'):
                        for sounds in particles:
                            if name == '':
                                if sounds.get('Sound') != None and sounds.get('Sound') != '':
                                    results[i][header[0]] = file
                                    results[i][header[1]] = particles.get('Name')
                                    for k in range(2, 7):
                                        results[i][header[k]] = sounds.get(header[k])
                                    i += 1
                            else:
                                if sounds.get('Sound') != None and name.lower() in sounds.get('Sound').split('\\')[-1].lower():
                                    results[i][header[0]] = file
                                    results[i][header[1]] = particles.get('Name')
                                    for k in range(2, 7):
                                        results[i][header[k]] = sounds.get(header[k])
                                    i += 1
    else:
        path = ''
        for file in dropedlist:
            if file.lower().endswith('.xml'):  # 특정 확장자만 열기
                xml = ET.parse(os.path.join(path, file))
                root = xml.getroot()

                try:
                    if root.iter('Particles') is not None: print('Particles - ' + file + ' 탐색중...')
                except:
                    pass


                for particles in root.iter('Particles'):
                    for sounds in particles:
                        if name == '':
                            if sounds.get('Sound') != None and sounds.get('Sound') != '':
                                results[i][header[0]] = file
                                results[i][header[1]] = particles.get('Name')
                                for k in range(2, 7):
                                    results[i][header[k]] = sounds.get(header[k])
                                i += 1
                        else:
                            if sounds.get('Sound') != None and name.lower() in sounds.get('Sound').split('\\')[
                                -1].lower():
                                results[i][header[0]] = file
                                results[i][header[1]] = particles.get('Name')
                                for k in range(2, 7):
                                    results[i][header[k]] = sounds.get(header[k])
                                i += 1

    return results, header


def sound_in_lyr(sound):

    header_SoundSpot = ['data', 'Layer', 'Name', 'Pos', 'EventType', 'bLoop', 'bOnce', 'bPlay', 'sndSource', 'InnerRadius', 'OuterRadius', 'iVolume']
    header_RA = ['data', 'Layer', 'Name', 'EventType', 'sndSound', 'bCentered', 'iVolume', 'iChanceOfOccuring', 'bDoNotOverlap', 'Shape']

    results_SoundSpot = defaultdict(dict)
    results_RA = defaultdict(dict)
    results_Shape = defaultdict(dict)
    i = 0
    s = 0
    r = 0

    if len(dropedlist) < 1:

        parents = os.path.abspath(os.path.join(os.path.abspath(os.pardir), ''))
        current = '80) Sound//02) Ambient'
        path = os.path.join(parents, current)

        for root, fir, files in os.walk(path):
            for file in files:
                if file.lower().endswith('.lyr'):
                    print('Layer - ' + file + ' 탐색중...')
                    xml = ET.parse(os.path.join(path, file))
                    root = xml.getroot()

                    for object in root.iter('Object'):

                        if object.get('Type') == 'Shape':
                            for ids in object.iter('Entities'):
                                idlist = []
                                for id in ids.iter('Entity'):
                                    results_Shape[i]['Name'] = object.get('Name')
                                    results_Shape[i]['Id'] = idlist
                                    idlist.append(id.get('Id'))
                                    i += 1

                    for object in root.iter('Object'):

                        if object.get('Type') == 'SoundSpot':
                            if sound == '':
                                results_SoundSpot[s][header_SoundSpot[0]] = file
                                for k in range(1, 5):
                                    results_SoundSpot[s][header_SoundSpot[k]] = object.get(header_SoundSpot[k])
                                for k in range(5, 12):
                                    results_SoundSpot[s][header_SoundSpot[k]] = object.find('Properties').get(
                                        header_SoundSpot[k])
                                s += 1
                            elif sound.lower() in object.find('Properties').get('sndSource').split('\\')[-1].lower():
                                results_SoundSpot[s][header_SoundSpot[0]] = file
                                for k in range(1, 5):
                                    results_SoundSpot[s][header_SoundSpot[k]] = object.get(header_SoundSpot[k])
                                for k in range(5, 12):
                                    results_SoundSpot[s][header_SoundSpot[k]] = object.find('Properties').get(
                                        header_SoundSpot[k])
                                s += 1

                        elif object.get('Type') == 'RandomAmbient':
                            if sound == '':

                                for v in object.iter():

                                    shapelist = []

                                    if v.tag.startswith('Sound') and v.attrib['sndSound'] != '':
                                        results_RA[r][header_RA[0]] = file

                                        for k in range(1, 4):
                                            results_RA[r][header_RA[k]] = object.get(header_RA[k])
                                            results_RA[r].update(v.attrib)

                                        for c in range(len(results_Shape)):
                                            if object.get('Id') in results_Shape[c]['Id']:
                                                shapelist.append(results_Shape[c]['Name'])

                                        results_RA[r]['Shape'] = shapelist

                                        r += 1

                            else:
                                for v in object.iter():

                                    shapelist = []

                                    if v.tag.startswith('Sound') and sound.lower() in v.attrib['sndSound'].split('\\')[-1].lower():
                                        results_RA[r][header_RA[0]] = file

                                        for k in range(1, 4):
                                            results_RA[r][header_RA[k]] = object.get(header_RA[k])
                                            results_RA[r].update(v.attrib)

                                        for c in range(len(results_Shape)):
                                            if object.get('Id') in results_Shape[c]['Id']:
                                                shapelist.append(results_Shape[c]['Name'])

                                        results_RA[r]['Shape'] = shapelist

                                        r += 1

    else:
        path = ''

        for file in dropedlist:
            if file.lower().endswith('.lyr'):
                print('Layer - ' + file + ' 탐색중...')
                xml = ET.parse(os.path.join(path, file))
                root = xml.getroot()

                for object in root.iter('Object'):

                    if object.get('Type') == 'Shape':
                        for ids in object.iter('Entities'):
                            idlist = []
                            for id in ids.iter('Entity'):
                                results_Shape[i]['Name'] = object.get('Name')
                                results_Shape[i]['Id'] = idlist
                                idlist.append(id.get('Id'))
                                i += 1

                for object in root.iter('Object'):

                    if object.get('Type') == 'SoundSpot':
                        if sound == '':
                            results_SoundSpot[s][header_SoundSpot[0]] = file
                            for k in range(1, 5):
                                results_SoundSpot[s][header_SoundSpot[k]] = object.get(header_SoundSpot[k])
                            for k in range(5, 12):
                                results_SoundSpot[s][header_SoundSpot[k]] = object.find('Properties').get(
                                    header_SoundSpot[k])
                            s += 1
                        elif sound.lower() in object.find('Properties').get('sndSource').split('\\')[
                            -1].lower():
                            results_SoundSpot[s][header_SoundSpot[0]] = file
                            for k in range(1, 5):
                                results_SoundSpot[s][header_SoundSpot[k]] = object.get(header_SoundSpot[k])
                            for k in range(5, 12):
                                results_SoundSpot[s][header_SoundSpot[k]] = object.find('Properties').get(
                                    header_SoundSpot[k])
                            s += 1

                    elif object.get('Type') == 'RandomAmbient':
                        if sound == '':

                            for v in object.iter():

                                shapelist = []

                                if v.tag.startswith('Sound') and v.attrib['sndSound'] != '':
                                    results_RA[r][header_RA[0]] = file

                                    for k in range(1, 4):
                                        results_RA[r][header_RA[k]] = object.get(header_RA[k])
                                        results_RA[r].update(v.attrib)

                                    for c in range(len(results_Shape)):
                                        if object.get('Id') in results_Shape[c]['Id']:
                                            shapelist.append(results_Shape[c]['Name'])

                                    results_RA[r]['Shape'] = shapelist

                                    r += 1

                        else:
                            for v in object.iter():

                                shapelist = []

                                if v.tag.startswith('Sound') and sound.lower() in \
                                        v.attrib['sndSound'].split('\\')[-1].lower():
                                    results_RA[r][header_RA[0]] = file

                                    for k in range(1, 4):
                                        results_RA[r][header_RA[k]] = object.get(header_RA[k])
                                        results_RA[r].update(v.attrib)

                                    for c in range(len(results_Shape)):
                                        if object.get('Id') in results_Shape[c]['Id']:
                                            shapelist.append(results_Shape[c]['Name'])

                                    results_RA[r]['Shape'] = shapelist

                                    r += 1

    return [results_SoundSpot, header_SoundSpot], [results_RA, header_RA]

class Form(QWidget):
    def __init__(self):
        super(Form, self).__init__()

        self.setWindowTitle("SoundSearch_AION")
        self.setMinimumSize(1300, 800)
        self.datas = []


        '''
        Layout_Base
        '''
        self.vb = QVBoxLayout()
        self.setLayout(self.vb)

        self.hbTop = QHBoxLayout()
        self.hbMid = QVBoxLayout()
        self.hbMidBot = QHBoxLayout()
        self.hbBot = QHBoxLayout()


        '''
        Groubbox
        '''
        self.gb_data = QGroupBox('Data Type')
        self.CB_Anim = QCheckBox("AnimationMarkers")
        self.CB_Particle = QCheckBox("Particles")
        self.CB_Layers = QCheckBox("Layers")
        self.CB_Anim.setChecked(True)
        self.CB_Particle.setChecked(True)
        self.CB_Layers.setChecked(True)
        self.ln = QLineEdit("")
        self.btn_name = QPushButton("Search")
        self.btn_clear = QPushButton("Clear")


        self.vbox = QHBoxLayout()
        self.vbox.addWidget(self.CB_Anim)
        self.vbox.addWidget(self.CB_Particle)
        self.vbox.addWidget(self.CB_Layers)
        self.vbox.addWidget(self.ln)
        self.vbox.addWidget(self.btn_name)
        self.vbox.addWidget(self.btn_clear)
        self.gb_data.setLayout(self.vbox)
        self.hbTop.addWidget(self.gb_data)

        self.vb.addLayout(self.hbTop)
        self.vb.addLayout(self.hbMid)
        self.vb.addLayout(self.hbMidBot)
        self.vb.addLayout(self.hbBot)

        '''
        Button
        '''

        self.btn_open = QPushButton("Open Selected Data File")
        self.btn_play = QPushButton("Open Selected Sound")

        self.btn_save = QPushButton("Save And Open xlsx")
        self.message = QMessageBox()

        self.hbMidBot.addWidget(self.btn_open)
        self.hbMidBot.addWidget(self.btn_play)
        self.hbBot.addWidget(self.btn_save)

        self.ln.returnPressed.connect(self.search)
        self.btn_name.clicked.connect(self.search)
        self.btn_clear.clicked.connect(self.clear)
        self.btn_save.clicked.connect(self.savefile)
        self.btn_play.clicked.connect(self.play)
        self.btn_open.clicked.connect(self.open)

        '''
        result tab
        '''
        self.tabs = QTabWidget()
        self.hbMid.addWidget(self.tabs)

        '''
        drop
        '''
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, 'Drop Files')
        self.tab1.layout = QVBoxLayout(self)
        # self.tab1.setLayout(self.tab1.layout)

        self.view = TestListView()
        self.view.fileDropped.connect(self.pictureDropped)
        self.tab1.layout.addWidget(self.view)
        self.view.addItem('탐색할 파일들을 끌어서 이곳에 놓아 주세요, 놓여진 파일이 없을시 AION_Data 하위의 모든 Sound 관련 XML 파일을 검사합니다.')


    def pictureDropped(self, l):
        self.view.clear()
        for url in l:
            if os.path.exists(url) and url.endswith('xml') or url.endswith('lyr') and not url in self.datas:
                item = QListWidgetItem(url, self.view)
                classfier.append(item.text())
                self.datas.append(item.text())
                item.setStatusTip(url)

    def clear(self):
        self.tabs.clear()
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, 'Drop Files')
        self.tab1.layout = QVBoxLayout(self)
        # self.tab1.setLayout(self.tab1.layout)

        classfier.remove()
        self.view = TestListView()
        self.view.fileDropped.connect(self.pictureDropped)
        self.tab1.layout.addWidget(self.view)
        self.view.addItem('탐색할 파일들을 끌어서 놓아 주세요, 놓여진 파일이 없을시 AION_Data 하위의 모든 Sound 관련 XML 파일을 검사합니다.')

        self.datas = []

    def addTab(self, name, dicts):
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, name)

        self.tab1.layout = QVBoxLayout(self)
        self.tab1.setLayout(self.tab1.layout)

        dict = dicts[0]
        header = dicts[1]

        self.tb_result = QTableWidget()
        self.tb_result.setAutoScroll(True)
        self.tb_result.showGrid()
        self.tb_result.clear()

        self.tb_result.setRowCount(len(dict))
        self.tb_result.setMinimumHeight(200)
        self.repaint()

        self.tb_result.setColumnCount(len(header))
        self.tb_result.setHorizontalHeaderLabels(header)


        for i in dict:
            for k in range(len(header)):
                item = QTableWidgetItem(str(dict[i][header[k]]))
                self.tb_result.setItem(i, k, item)

        self.tb_result.setEditTriggers(QTableWidget.NoEditTriggers)

        remain = len(header)

        for i in range(1, len(header)):

            if 'Sound' == header[i] or 'file' in header[i] or 'snd' in header[i]:
                self.tb_result.setColumnWidth(i, 370)
                remain -= 1
            elif header[i].startswith('i') or header[i].startswith('b') or header[i].startswith('Event') or 'Radius' in header[i] or header[i].startswith('Sound') or header[i].startswith('vol'):
                self.tb_result.setColumnWidth(i, 60)
                remain -= 1
            else:
                self.tb_result.setColumnWidth(i, 1100 // remain)
            if len(dropedlist) > 0:
                self.tb_result.setColumnWidth(0, 320)
            else:
                self.tb_result.setColumnWidth(0, 200)

        self.tb_result.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tab1.layout.addWidget(self.tb_result)

        return


    def search(self):

        self.tabs.clear()

        if self.CB_Anim.isChecked():
            anim = sound_in_anim(self.ln.text())
            if len(anim[0]) != 0:
                print('AnimationMarkers - 결과 저장 중...')
                self.addTab('AnimationMarkers', anim)

        if self.CB_Particle.isChecked():
            particle = sound_in_particles(self.ln.text())
            if len(particle[0]) != 0:
                print('Particles - 결과 저장 중...')
                self.addTab('Particles', particle)

        if self.CB_Layers.isChecked():
            layer = sound_in_lyr(self.ln.text())
            if len(layer[0][0]) != 0:
                print('Layer - 결과 저장 중...')
                self.addTab('Layer_Spot', layer[0])
            if len(layer[1][0]) != 0:
                self.addTab('Layer_RA', layer[1])

        print('완료')
        return

    def play(self):
        self.playsound()
        return

    def open(self):
        self.openjson()
        return


    def playsound(self):

        try:
            result = self.tabs.currentWidget().findChildren(QTableWidget)[0]

            parents = os.path.abspath(os.path.join(os.path.abspath(os.pardir), os.pardir))
            current = 'AION_Data//'
            path = os.path.join(parents, current)

            for i in range(result.columnCount()):
                if result.item(result.currentRow(), i).text().endswith('ogg'):
                    os.startfile(path + '\\' + result.item(result.currentRow(), i).text())


        except AttributeError:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_AION")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("실행할 파일을 선택해 주세요.")
            self.message0.exec()

        except IndexError:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_AION")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("실행할 파일을 선택해 주세요.")
            self.message0.exec()

        except OSError:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_AION")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("지정된 파일 열기 위한 응용 프로그램이 설정 되어 있지 않습니다.")
            self.message0.exec()

        return

    def openjson(self):
        try:

            result = self.tabs.currentWidget().findChildren(QTableWidget)[0]

            if self.tabs.tabText(self.tabs.currentIndex()) == 'AnimationMarkers':
                parents = os.path.abspath(os.path.join(os.path.abspath(os.pardir), os.pardir))
                current = 'AION_Data//Data//animationmarkers'
                path = os.path.join(parents, current)

            elif self.tabs.tabText(self.tabs.currentIndex()) == 'Particles':
                parents = os.path.abspath(os.path.join(os.path.abspath(os.pardir), os.pardir))
                current = 'AION_Data//Editor//particles'
                path = os.path.join(parents, current)

            else:
                parents = os.path.abspath(os.path.join(os.path.abspath(os.pardir), ''))
                current = '80) Sound//02) Ambient'
                path = os.path.join(parents, current)

            if len(dropedlist) < 1:
                os.startfile(path + '\\' + result.item(result.currentRow(), 0).text())
            else:
                os.startfile(result.item(result.currentRow(), 0).text())

        except AttributeError:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_AION")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("실행할 파일을 선택해 주세요.")
            self.message0.exec()

        except IndexError:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_AION")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("실행할 파일을 선택해 주세요.")
            self.message0.exec()

        except OSError:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_AION")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("지정된 파일 열기 위한 응용 프로그램이 설정 되어 있지 않습니다.")
            self.message0.exec()

        return


    def savefile(self):



        filename = QFileDialog.getSaveFileName(self, 'Save file', "", ".xlsx(*.xlsx)")

        print(filename + ' 저장 중...')

        wb = Workbook()


        for i in range(self.tabs.count()):

            self.tabs.setCurrentIndex(i)
            tab = self.tabs.currentWidget()
            result = tab.findChildren(QTableWidget)[0]

            ws = wb.create_sheet(self.tabs.tabText(i))


            header = []

            for i in range(result.columnCount()):
                header.append(result.horizontalHeaderItem(i).text())

            rowindex = 1
            colindex = 1

            for h in header:
                ws.cell(row=rowindex, column=colindex).value = h
                ws.cell(row=rowindex, column=colindex).font = Font(bold=True, color='ffffff')
                ws.cell(row=rowindex, column=colindex).fill = PatternFill("solid", fgColor="404040")
                colindex += 1

            ws.freeze_panes = 'A2'

            for i in range(1, len(header)+1):
                if 'Sound' == header[i-1] or 'file' in header[i-1] or 'snd' in header[i-1]:
                    ws.column_dimensions[get_column_letter(i)].width = 60
                elif header[i-1].startswith('b') or header[i-1].startswith('i') or header[i-1].startswith('Event'):
                    ws.column_dimensions[get_column_letter(i)].width = 7
                else:
                    ws.column_dimensions[get_column_letter(i)].width = 20


            ws.auto_filter.ref = "A1:" + get_column_letter(len(header)) + "1"


            for currentColumn in range(1, result.columnCount()+1):
                for currentRow in range(1, result.rowCount()+1):
                    try:
                        teext = str(result.item(currentRow-1, currentColumn-1).text())
                        ws.cell(currentRow+1, currentColumn).value = teext

                    except AttributeError:
                        pass

            # wb.remove(wb['Sheet'])
        self.tabs.setCurrentIndex(0)
        wb.remove(wb['Sheet'])

        try:
            wb.save(filename[0])
            self.message1 = QMessageBox()
            self.message1.setWindowTitle("SoundSearch_AION")
            self.message1.setText(str(filename[0]) + '\n' + '\n' + "저장 완료!")
            self.message1.exec()
            os.startfile(str(filename[0]))
        except PermissionError:
            self.message1 = QMessageBox()
            self.message1.setWindowTitle("SoundSearch_AION")
            self.message1.setText("저장 권한이 없습니다. 엑셀이 닫혀있는지 확인해 주세요.")
            self.message1.exec()


        wb.close()

        return

class TestListView(QListWidget):

    fileDropped = QtCore.Signal(list)

    def __init__(self, parent=None):
        super(TestListView, self).__init__(parent)
        self.setAcceptDrops(True)
        self.setIconSize(QtCore.QSize(72, 72))

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls:
            event.accept()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls:
            event.setDropAction(QtCore.Qt.CopyAction)
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls:
            event.setDropAction(QtCore.Qt.CopyAction)
            event.accept()
            links = []
            for url in event.mimeData().urls():
                links.append(str(url.toLocalFile()))
            self.fileDropped.emit(links)
        else:
            event.ignore()





app = QApplication([])
GUI = Form()
GUI.show()
app.exec_()
