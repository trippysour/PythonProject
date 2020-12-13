import sys
from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font


###############################################사용자 입력#################################################


#사용자 입력 (.txt 파일)
print("Name of the txt file?")
input_filename = str(input())


#입력 오류 안전장치
if input_filename.endswith(".txt"):
	file_txt = input_filename
	print("Continuing process...")
elif "." in input_filename:
	print("Only .txt format is supported. Please re-run the program and try again.")
	print("Press any button to close the program...")
	sys.exit(0)
else:
	file_txt = input_filename + ".txt"
	print("Continuing process...")


################################################시트 설정##################################################


#'object sheet' 클래스 생성 (sheet var name, sheet title, header set, rowname) 근데 이게 더 코드 길어질듯

#엑셀 파일 생성
file_xl = Workbook()

#Sheet1 설정 - 시트 이름, 칼럼 구성
sheet1 = file_xl.active
sheet1.title = "SpotSounds"

s1_header = {
	'Layer': 1,
	'Name': 2,
	'bEnabled': 3,
	'bLoop': 4,
	'InnerRadius': 5,
	'OuterRadius': 6,
	'iVolume': 7,
	'sndSource': 8,
	}

for i in s1_header:
	sheet1.cell(row = 1, column = s1_header.get(i), value = i)

#Sheet2 설정 - 시트 이름, 칼럼 구성 ****EntityId 지우지 말아주세요
sheet2 = file_xl.create_sheet()
sheet2.title = "Shapes"

s2_header = {
	'Layer': 1,
	'Name': 2,
	'GroupId': 3,
	'RandomAmbient': 4,
	'EntityId': 5
}

for i in s2_header:
	sheet2.cell(row = 1, column = s2_header.get(i), value = i)

#Sheet3 설정 - 시트 이름, 칼럼 구성 ****Id 지우지 말아주세요
sheet3 = file_xl.create_sheet()
sheet3.title = "RandomAmbient"

s3_header = {
	'Layer': 1,
	'Name': 2,
	'Variations': 3,
	'bCentered': 4,
	'bDoNotOverlap': 5,
	'iChanceOfOccuring': 6,
	'iVolume': 7,
	'sndSource': 8,
	'Id': 9
}

for i in s3_header:
	sheet3.cell(row = 1, column = s3_header.get(i), value = i)


###############################################함수 정의##################################################


#반복작업 묶음 - Split은 텍스트마다 기준이 달라서 함수 부르기 전에 먼저 수행해야함
def ParseSplitline(sheet, headerset, rowname):
	#look for properties in each splitline
	for splitline in splitlines:
		splitline = splitline.replace('"','')

		for i in headerset:
			if splitline.find(i + "=") != -1:
				#On finding a property, write value to the according cell
				try:
					sheet.cell(row = rowname, column = headerset.get(i)).value = float(splitline[len(i)+1:].replace('/>',''))
				except:
					sheet.cell(row = rowname, column = headerset.get(i)).value = splitline[len(i)+1:].replace('/>','')


################################################파싱 시작#################################################


#open txt file, read lines, and close txt file
with open(file_txt, mode="r", encoding="utf-8") as file:
	lines = file.readlines()
file.close()

r1, r2, r3 = 1, 2, 2

for line in lines:

	#Sheet1 내용 채워넣기
	if line.find(" Type=\"SoundSpot\" ") != -1:
		r1 = r1 + 1		#add row index, increment per object only
		splitlines = line.split(' ')
		ParseSplitline(sheet1, s1_header, r1)

	if line.find("<Properties ") != -1:
		splitlines = line.split(' ')
		ParseSplitline(sheet1, s1_header, r1)


	#Sheet2 내용 채워넣기
	if line.find(" Type=\"Shape\" ") != -1:
#		r2 = r2 + 1		#add row index, increment per object only
		splitlines = line.split(' ')
		ParseSplitline(sheet2, s2_header, r2)
	#Entity Id는 조금 특수한 관계로 별도로 처리합니다.
	if line.find("<Entity Id=") != -1:
		sheet2.cell(row = r2, column = s2_header.get('EntityId')).value = line.strip()[11:].replace('"','').replace('/>','')

		r2 = r2 + 1


	#Sheet3 내용 채워넣기
	if line.find(" Type=\"RandomAmbient\" ") != -1:
#		r3 = r3 + 1		#add row index, increment per object onl
		splitlines = line.split(' ')
		ParseSplitline(sheet3, s3_header, r3)
	#Look for Sound variations that are not empty, and split line to get properties
	if line.find("<Sound") != -1:
		if line.find("bCentered=\"0\" bDoNotOverlap=\"1\" iChanceOfOccuring=\"0\" iVolume=\"0\" sndSound=\"\"") != -1:
			pass
		else:
			line = line.replace('"','')
			splitlines = line.split(' ')
			ParseSplitline(sheet3, s3_header, r3)

			#Variation은 예외 규칙을 사용하므로 별도로 처리, sndSource는 sndSound라고 오타난 채로 사용하고 계셔서 별도로 처리
			sheet3.cell(row = r3, column = s3_header.get('Variations')).value = splitlines[0].replace('<','')
			sheet3.cell(row = r3, column = s3_header.get('sndSource')).value = splitlines[5][9:].replace('/>','')

			r3 = r3 + 1	


################################Shape-RandomAmbient 레퍼런스 작업##########################################


for i in range(2, r3+1):
	#See if cell has Id info	#If it does, see which Shape has matching Entity Id
	if sheet3.cell(row = i, column = s3_header.get('Id')).value != None:
		for j in range(2, r2+1):
			#If Amb Id matches Shape Entity Id, write down RandomAmbient info to both Sheet2
			if sheet2.cell(row = j, column = s2_header.get('EntityId')).value == sheet3.cell(row = i, column = s3_header.get('Id')).value:
				sheet2.cell(row = j, column = s2_header.get('RandomAmbient')).value = sheet3.cell(row = i, column = s3_header.get('Name')).value


################################################서식 설정##################################################

#사용자에게 불필요한 column 제거 (Id, Entity Id)
sheet2.delete_cols(s2_header.get('EntityId'))
sheet3.delete_cols(s3_header.get('Id'))

def styleSheet(sheet, headerset):

	#add filter
	sheet.auto_filter.ref = sheet.dimensions

	#autowidth
	dims = {}
	for row in sheet.rows:
		for cell in row:
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
	for col, value in dims.items():
		sheet.column_dimensions[col].width = value

	#style
	for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headerset)):
		for cell in rows:
			if cell.value != None:
				cell.fill = PatternFill("solid", fgColor="404040")
				cell.font = Font(bold=True, color="FFFFFF")
				cell.alignment = Alignment(horizontal="center")	


styleSheet(sheet1, s1_header)
styleSheet(sheet2, s2_header)
styleSheet(sheet3, s3_header)


################################################파일 저장##################################################

#처음에 입력한 텍스트 파일과 같은 이름으로 엑셀 파일 생성
file_xl.save(filename = file_txt.replace('.txt','') + ".xlsx")

print("Request completed. Please press Enter to close the program.")
input("prompt: ")
