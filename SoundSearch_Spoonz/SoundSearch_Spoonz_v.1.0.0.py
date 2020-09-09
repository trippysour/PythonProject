import os
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from waapi import WaapiClient, CannotConnectToWaapiException
from jsonpath_ng import parse
from collections import defaultdict
from PySide2.QtWidgets import QWidget, QVBoxLayout, QMessageBox, QHBoxLayout, QTableWidget, QLineEdit, QPushButton, QApplication, QLabel, QTableWidgetItem, QFileDialog, QAbstractItemView
from PySide2.QtGui import QColor

parents = '..\\Assets\\Resources\\Outgame\\Data\\Sound' # 실제 path
#parents = os.getcwd()  # 개발전용
os.chdir(parents)  # json 폴더 지정

header = ['file', 'ContentsKey', 'animName', 'soundName', 'targetObjName', 'sequenceTime', 'playOneShot', 'dontDestroy']

def jsontodict(sound):
    results = defaultdict(dict)
    i = 0

    for root, dirs, files in os.walk('./'):
        for file in files:
            if file.lower().endswith('.json'):  # 특정 확장자만 열기
                fpath = os.path.join(root, file)  # join 으로 합쳐야지 str이 아닌 dir로 인식
                with open(fpath, encoding='cp949') as json_file:  # 인코딩 지정하지 않으면 에러 발
                    json_data = json.load(json_file)
                    json_str = json.dumps(json_data)

                    for key in parse('$..ContentsKey').find(json.loads(json_str)):
                        data = key.context.value['m_saveDataList']
                        for match in parse('$..soundName').find(data):
                            if sound == '':  # 빈 칸이면 모두
                                results[i][header[0]] = file
                                results[i][header[1]] = key.context.value[header[1]]
                                results[i].update(match.context.value)
                                i += 1
                            elif sound.lower() in match.value.lower():  # 특정 단어 포함하는지, 대소문자 구분 없이 하기 위해 둘다 소문자 처리해서 비교
                                results[i][header[0]] = file
                                results[i][header[1]] = key.context.value[header[1]]
                                results[i].update(match.context.value)
                                i += 1

    return results


class Form(QWidget):
    def __init__(self):
        super(Form, self).__init__()
        self.setWindowTitle("SoundSearch_Spoonz")
        self.setMinimumSize(875, 600)

        self.vb = QVBoxLayout()
        self.setLayout(self.vb)
        self.hbTop = QHBoxLayout()
        self.hbMid = QVBoxLayout()
        self.hbMidBot = QHBoxLayout()
        self.hbBot = QHBoxLayout()
        self.vb.addLayout(self.hbTop)
        self.vb.addLayout(self.hbMid)
        self.vb.addLayout(self.hbMidBot)
        self.vb.addLayout(self.hbBot)

        self.ln = QLineEdit("검색 할 Event 명에 들어간 단어를 입력해 주세요")
        self.btn_name = QPushButton("Search")
        self.btn_all = QPushButton("Search All")
        self.btn_open = QPushButton("Open Selected Json")
        self.btn_play = QPushButton("Post Selected Event")
        self.btn_stop = QPushButton("Stop All")
        self.lb_result = QLabel("결과 :")
        self.tb_result = QTableWidget()
        self.tb_result.setAutoScroll(True)
        self.tb_result.showGrid()

        self.btn_check = QPushButton("Check Events in Wwise")
        self.btn_save = QPushButton("Save And Open xlsx")
        self.message = QMessageBox()

        self.hbTop.addWidget(self.ln)
        self.hbTop.addWidget(self.btn_name)
        self.hbTop.addWidget(self.btn_all)
        self.hbMid.addWidget(self.lb_result)
        self.hbMid.addWidget(self.tb_result)
        self.hbMidBot.addWidget(self.btn_open)
        self.hbMidBot.addWidget(self.btn_play)
        self.hbMidBot.addWidget(self.btn_stop)
        self.hbMidBot.addWidget(self.btn_check)
        self.hbBot.addWidget(self.btn_save)

        self.ln.returnPressed.connect(self.search)
        self.btn_name.clicked.connect(self.search)
        self.btn_all.clicked.connect(self.search_all)
        self.btn_save.clicked.connect(self.savefile)
        self.btn_play.clicked.connect(self.play)
        self.btn_play.clicked.connect(self.stop)
        self.btn_open.clicked.connect(self.open)
        self.btn_check.clicked.connect(self.check)

    def search_all(self):
        self.showresult(jsontodict(''))
        return

    def search(self):
        self.showresult(jsontodict(self.ln.text()))
        return

    def play(self):
        self.postevent()
        return

    def open(self):
        self.openjson()
        return

    def check(self):
        self.checkevents()

    def stop(self):
        self.stopall()


    def checkevents(self):
        try:
            client = WaapiClient()

            notfounds = []

            for i in range(self.tb_result.rowCount()):
                self.tb_result.item(i, 3).setBackgroundColor(QColor(255, 255, 255))
                event = self.tb_result.item(i, 3).text()

                arg = {
                    "from": {
                        "search": [
                            event
                        ],
                    },
                    "transform": [
                        {
                            "where": [
                                "type:isIn",
                                [
                                    "Event"
                                ]
                            ],
                            "where": [
                                "name:matches", event
                            ]
                        }
                    ]
                }
                options = {
                    "return": ['name']
                }

                if not client.call("ak.wwise.core.object.get", arg, options=options)['return']:
                    notfounds.append(event)
                    self.tb_result.item(i, 3).setBackgroundColor(QColor(255,0,0))
                    self.tb_result.repaint()

                elif event != client.call("ak.wwise.core.object.get", arg, options=options)['return'][0]['name']:
                    notfounds.append(event)
                    self.tb_result.item(i, 3).setBackgroundColor(QColor(255, 0, 0))
                    self.tb_result.repaint()

            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_Spoonz")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("총 " + str(len(notfounds)) + "개의 Event를 찾을 수 없습니다.")
            self.message0.exec()


        except CannotConnectToWaapiException:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_Spoonz")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("WAAPI에 연결하지 못했습니다. : Wwise가 켜져있고 WAAPI가 Enabled 되어 있는지 체크해 주세요.")
            self.message0.exec()


    def showresult(self, dict):

        self.tb_result.clear() # 채우기 전에 초기화

        self.tb_result.setRowCount(len(dict))
        self.lb_result.setText("결과 : 총 " + str(len(dict)) + " 개의 Event를 찾았습니다.")
        self.repaint() # 이걸 해줘야 레이블이 업데이트 됨

        self.tb_result.setColumnCount(len(header))
        self.tb_result.setHorizontalHeaderLabels(header)

        for i in dict:
            for k in range(len(header)):
                item = QTableWidgetItem(str(dict[i][header[k]]))
                self.tb_result.setItem(i, k, item)

        # len(dict) == i 는 row, 행의 갯수
        # len(dict[0]) == len(header) == k 는 col, 열의 갯수

        self.tb_result.setEditTriggers(QTableWidget.NoEditTriggers) # 에디팅 막음
        self.tb_result.setSelectionMode(QAbstractItemView.SingleSelection) # 중복선택 불가능 하게

        return


    def postevent(self):

        try:
            client = WaapiClient()

            try:
                event = self.tb_result.item(self.tb_result.currentRow(), 3).text()

                arg = {
                    "from": {
                        "search": [
                            event
                        ],
                    },
                    "transform": [
                        {
                            "where": [
                                "type:isIn",
                                [
                                    "Event"
                                ]
                            ],
                            "where": [
                                "name:matches", event
                            ]
                        }
                    ]
                }
                result = client.call("ak.wwise.core.object.get", arg)

                try:
                    if result['return'][0]['name'] == event:
                        arg = {
                            "event": result['return'][0]['name'],
                            "gameObject": 18446744073709551614  # Transport ID
                        }

                        client.call("ak.wwise.ui.bringToForeground")
                        client.call("ak.soundengine.postEvent", arg)

                    else:
                        self.message0 = QMessageBox()
                        self.message0.setWindowTitle("SoundSearch_Spoonz")
                        self.message0.setIcon(QMessageBox.Warning)
                        self.message0.setText("Event를 찾을 수 없습니다.")
                        self.message0.exec()


                except IndexError:
                    self.message0 = QMessageBox()
                    self.message0.setWindowTitle("SoundSearch_Spoonz")
                    self.message0.setIcon(QMessageBox.Warning)
                    self.message0.setText("Event를 찾을 수 없습니다.")
                    self.message0.exec()


            except AttributeError:
                self.message0 = QMessageBox()
                self.message0.setWindowTitle("SoundSearch_Spoonz")
                self.message0.setIcon(QMessageBox.Warning)
                self.message0.setText("재생할 Event를 선택해 주세요.")
                self.message0.exec()


        except CannotConnectToWaapiException:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_Spoonz")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("WAAPI에 연결하지 못했습니다. : Wwise가 켜져있고 WAAPI가 Enabled 되어 있는지 체크해 주세요.")
            self.message0.exec()

    def stopall(self):
        try:
            client = WaapiClient()

            arg = {
                "gameObject": 18446744073709551614  # Transport ID
            }

            client.call("ak.soundengine.stopAll", arg)

        except CannotConnectToWaapiException:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_Spoonz")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("WAAPI에 연결하지 못했습니다. : Wwise가 켜져있고 WAAPI가 Enabled 되어 있는지 체크해 주세요.")
            self.message0.exec()


    def openjson(self): #아무 응답이 없을 때는 연결프로그램 확인

        try:
            os.startfile(parents + '/' + self.tb_result.item(self.tb_result.currentRow(), 0).text())

        except AttributeError:
            self.message0 = QMessageBox()
            self.message0.setWindowTitle("SoundSearch_Spoonz")
            self.message0.setIcon(QMessageBox.Warning)
            self.message0.setText("실행할 Json을 선택해 주세요.")
            self.message0.exec()
        return

    def savefile(self):
        filename = QFileDialog.getSaveFileName(self, 'Save file', "", ".xlsx(*.xlsx)")

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sound'

        rowindex = 1
        colindex = 1

        for h in header:
            ws.cell(row=rowindex, column=colindex).value = h
            ws.cell(row=rowindex, column=colindex).font = Font(bold=True, color='ffffff')
            ws.cell(row=rowindex, column=colindex).fill = PatternFill("solid", fgColor="404040")
            colindex += 1

        ws.freeze_panes = 'A2'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30

        ws.auto_filter.ref = "A1:H1"


        for currentColumn in range(1, self.tb_result.columnCount()+1):
            for currentRow in range(1, self.tb_result.rowCount()+1):
                try:
                    teext = str(self.tb_result.item(currentRow-1, currentColumn-1).text())
                    ws.cell(currentRow+1, currentColumn).value = teext

                except AttributeError:
                    pass


        try:
            wb.save(filename[0])
            self.message1 = QMessageBox()
            self.message1.setWindowTitle("SoundSearch_Spoonz")
            self.message1.setText(str(filename[0]) + '\n' + '\n' + "저장 완료!")
            self.message1.exec()
            os.startfile(str(filename[0]))
        except PermissionError:
            self.message1 = QMessageBox()
            self.message1.setWindowTitle("SoundSearch_Spoonz")
            self.message1.setText("저장 권한이 없습니다. 엑셀이 닫혀있는지 확인해 주세요.")
            self.message1.exec()




        wb.close()

        return

app = QApplication([])
GUI = Form()

GUI.show()
app.exec_()
